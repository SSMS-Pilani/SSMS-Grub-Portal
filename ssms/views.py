import math
from datetime import datetime, date, timedelta

import openpyxl
import xlsxwriter
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
# To be triggered
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponseRedirect
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views.decorators.cache import cache_control
from django.contrib.auth.decorators import user_passes_test

from ssms.forms import GrubForm, Grub_CoordUserForm, Grub_CoordUserProfileForm, ExcelUpload, VegForm, NonVegForm, \
    BothForm, CoordStudentRegForm, GrubFormEdit, UploadFileForm, FeedbackForm
from ssms.models import Grub, Grub_Coord, Grub_Student, Veg, NonVeg, Both, Student, DateMailStatus, Grub_Member, \
    Grub_Invalid_Students, Batch, Meal


def datechecker(gmid):
    grub = Grub.objects.get(gm_id=gmid)
    current_date = date.today()
    if grub.deadline2 > current_date:
        return 2  # coord reg open
    elif grub.deadline2 <= current_date <= grub.deadline:
        return 1  # student register/cancel  open, coord reg closed
    elif grub.deadline < current_date <= grub.date:
        return 3  # student reg/cancel closed
    elif grub.date < current_date:
        return 4  # coord spot signing upload


def home(request):
    return render(request, 'ssms/home.html')


def about(request):
    return render(request, 'ssms/about.html')


def contact(request):
    return render(request, 'ssms/contact.html')


def send(request):  # cancellation mail
    if request.user.is_superuser:
        grub = Grub.objects.filter(status="Active", mails="Not Sent")
        current_date = date.today()
        diff = timedelta(days=1)
        final = current_date + diff
        sent_to_students = []
        for i in grub:
            diff = datetime.strptime(str(i.date), '%Y-%m-%d')
            e = date.strftime(diff, "%d %B %Y")
            v = datetime.strptime(str(i.deadline), '%Y-%m-%d')
            h = date.strftime(v, "%d %B %Y")
            if current_date == i.deadline2 or final == i.deadline2:
                all_students = Grub_Student.objects.filter(gm_id=i.gm_id, status="Signed Up")
                for q in range(len(all_students) // 99 + 1):  # send mails in batches of 100
                    students = all_students[q * 99:(q + 1) * 99]
                    student_id_list = [str(student.user_id) + "@pilani.bits-pilani.ac.in" for student in students]
                    subject, from_email = str(i.name), 'ssms.pilani@gmail.com'
                    text_content = 'This is an important message.'
                    html_content = "<body><p>This is to inform you that you have been signed up for the <strong> " \
                                   + str(i.name) + "</strong> that is to take place on <strong>" + e + "</strong> </p> <p>In case you wish to cancel your signing,\
					please visit <a href=http://www.ssms-pilani.org/ssms/student/grub/" + str(i.gm_id) + "/ >SSMS Grub Portal</a>, \
					before 12 midnight,<strong>" + h + "</strong>. Any requests made after the deadline will not be entertained. </p> <p><strong> If the above url doesn't work, please click \
					<a href=https://ssmsbitspilani.herokuapp.com/ssms/student/grub/" + str(i.gm_id) + "/ >here</a> </strong></p> <p>Thank you.</p><p>Grub Committee, SSMS</p></body>"
                    msg = EmailMultiAlternatives(
                        subject,
                        text_content,
                        from_email,
                        cc=student_id_list,
                        bcc=["f2014623@pilani.bits-pilani.ac.in", "f2015040@pilani.bits-pilani.ac.in"]
                    )
                    msg.attach_alternative(html_content, "text/html")
                    msg.send(fail_silently=False)
                    sent_to_students.append("Sent mail for " + str(i.name) + " to " + str(len(student_id_list)) + str(student_id_list))
                    for j in students:
                        j.mail = "Sent"
                        j.save()
                i.mails = "Sent"
                i.save()

        return HttpResponse("Sent python mail" + str(sent_to_students))
    else:
        return HttpResponseRedirect("/ssms")


def send2(request):
    if request.user.is_superuser:
        grubs = Grub.objects.filter(status="Active", mails="Sent")
        today = date.today()
        date_difference = timedelta(days=1)
        next_day = today + date_difference
        sent_to_students = []
        for grub in grubs:
            if grub.meal == "Veg":
                veg = Veg.objects.get(gm_id=grub.gm_id)
                meal = str(veg.v_venue)
            elif grub.meal == "Non Veg":
                non_veg = NonVeg.objects.get(gm_id=grub.gm_id)
                meal = str(non_veg.n_venue)
            else:
                veg = Veg.objects.get(gm_id=grub.gm_id)
                non_veg = NonVeg.objects.get(gm_id=grub.gm_id)
                meal = str(veg.v_venue) + " and " + str(non_veg.n_venue)
            grub_date = date.strftime(datetime.strptime(str(grub.date), '%Y-%m-%d'), "%d %B %Y")
            if today == grub.date or next_day == grub.date:
                all_students = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up")
                k = len(all_students) // 99
                for q in range(k + 1):
                    students = all_students[q * 99:(q + 1) * 99]
                    students_id_list = map(lambda x: str(x.user_id) + "@pilani.bits-pilani.ac.in", students)
                    subject, from_email = str(grub.name) + " (Reminder)", 'ssms.pilani@gmail.com'
                    text_content = 'This is an important message.'
                    html_content = "<body><p>This is to remind you that you that you have been signed up for <strong> " + \
                                   str(grub.name) + "</strong> which will take place on <strong>" + grub_date + "</strong> at the <strong>" + meal + \
                                   "</strong> Mess. </p> <p>Wristbands for the same are available at your mess counter, and you are requested to collect the same if you haven't already.</p><strong><p>Entry into the grub shall not be allowed if you are not wearing the wristband.</p></strong><p>Limited on spot signings will be available. Please carry your ID cards for the same. </p><p>Thank you.</p><p>Grub Committee, SSMS</p></body>"
                    msg = EmailMultiAlternatives(subject, text_content, from_email, cc=students_id_list, bcc=["f2014623@pilani.bits-pilani.ac.in", "f2015040@pilani.bits-pilani.ac.in"])
                    msg.attach_alternative(html_content, "text/html")
                    msg.send(fail_silently = False)
                    sent_to_students.append("Sent mail for " + str(grub.name) + " to " + str(len(students_id_list)) + str(students_id_list))
                    for j in students:
                        j.mail = "Sent"
                        j.save()
                grub.mails = "Sent2"
                grub.save()

        return HttpResponse("Sent python mail" + str(sent_to_students))
    else:
        return HttpResponseRedirect("/ssms")


def ssms_grub_sendmail1(request, gmid):
    grubid = request.GET.get('grubid')
    datemail = DateMailStatus.objects.get(date=datetime.now())
    try:
        grub = Grub.objects.get(name=str(grubid))
        print(gmid)
        d = datetime.strptime(str(grub.date), '%Y-%m-%d')
        e = date.strftime(d, "%d %B %Y")
        v = datetime.strptime(str(grub.deadline), '%Y-%m-%d')
        h = date.strftime(v, "%d %B %Y")
        count = 0
        abcd = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up", mail="Not Sent")
        k = len(abcd) // 99
        for q in range(k + 1):
            a = []
            students = abcd[q * 99:(q + 1) * 99]
            for j in students:
                a.append(str(j.user_id) + "@pilani.bits-pilani.ac.in")
                j.mail = "Sent"
                j.save()
            subject, from_email = str(grub.name) + " | " + e, 'ssms.pilani@gmail.com'
            text_content = 'This is an important message.'
            html_content = "<body><p>This is to inform you that you have been signed up for the <strong> " + str(
                grub.name) + "</strong> that is to take place on <strong>" + e + "</strong> </p> <p>In case you wish to cancel your signing, please visit <a href=http://grub.ssms-pilani.org/ssms/student/grub/" + str(
                grub.gm_id) + "/ >SSMS Grub Portal</a>, before 12 midnight,<strong>" + h + "</strong>. Any requests made after the deadline will not be entertained. </p><p><strong> If the above url doesn't work, please click <a href=https://ssmsbitspilani.herokuapp.com/ssms/student/grub/" + str(
                grub.gm_id) + "/ >here</a></strong></p><p>Thank you.</p><p>Grub Committee, SSMS</p></body>"
            msg = EmailMultiAlternatives(subject, text_content, from_email, cc=a, bcc=[
                "f2015040@pilani.bits-pilani.ac.in"])
            msg.attach_alternative(html_content, "text/html")
            print(a)
            try:
                msg.send(fail_silently=False)
                count = count + len(a)
                datemail.mails = datemail.mails + len(a)
                datemail.save()
            except Exception as e:
                print(e)
                for j in students:
                    j.mail = "Not Sent"
                    j.save()
                left = len(abcd) - count
                data = {'is_taken': "Only " + str(count) + " mails were sent succesfully. " +
                                    str(left) + " mails are left to be send. Error " + str(e)}
                return JsonResponse(data)
        grub.mails = "Sent"
        grub.save()
        data = {'is_taken': "All the mails (" + str(len(abcd)) + ") were sent successfully"}
        return JsonResponse(data)
    except Exception:
        print("here")
        data = {'is_taken': "Internal Error"}
        return JsonResponse(data)


def ssms_grub_sendmail3(grub, forloop, datemail, d, e, getspotsigning):
    count = 0
    data = ""
    allstu = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up")
    try:
        if (forloop == 1 or forloop == 3):
            if (forloop == 1):
                veg = Veg.objects.get(gm_id=grub)
                venue = veg.v_venue
            else:
                veg = Both.objects.get(gm_id=grub)
                venue = veg.veg_venue
            abcd = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up", meal="Veg")
            k = len(abcd) // 99
            for q in range(k + 1):
                a = []
                students = abcd[q * 99:(q + 1) * 99]
                for j in students:
                    if (j.mail != "Sent2"):
                        a.append(str(j.user_id) + "@pilani.bits-pilani.ac.in")
                        j.mail = "Sent2"
                        j.save()
                if (len(a) > 0):
                    subject, from_email = str(grub.name) + " | " + e + \
                                          " | Veg | " + venue, 'ssms.pilani@gmail.com'
                    text_content = 'This is an important message.'
                    html_content = "<body><p>This is to remind you that you that you have been signed up for <strong> " + \
                                   str(
                                       grub.name) + "</strong> that is going to take place on <strong>" + e + "</strong> at the " + venue + ".\
					You are required \
					to collect your stubs from the PitStop counter in your mess during meal timings today.</p>\
	<strong><p>Entry into the grub shall not be allowed if you are not wearing the wristband.</p></strong>\
					<p>" + getspotsigning + " </p><p>Thank you.</p>\
					<p>Regards,</p>\
					<p>Grub Committee, SSMS</p></body>"
                    # print a
                    msg = EmailMultiAlternatives(subject, text_content, from_email, cc=a, bcc=[
                        "f2015040@pilani.bits-pilani.ac.in"])
                    msg.attach_alternative(html_content, "text/html")
                    try:
                        msg.send(fail_silently=False)
                        count = count + len(a)
                        datemail.mails = datemail.mails + len(a)
                        datemail.save()
                    except Exception as e:
                        for j in students:
                            j.mail = "Sent"
                            j.save()
                        left = len(allstu) - count
                        data = "Only " + str(count) + " -Veg- mails were sent succesfully. " + \
                               str(left) + " mails are left to be send. Error " + str(e)
                        return data
        if (forloop == 2 or forloop == 3):
            if (forloop == 1):
                veg = NonVeg.objects.get(gm_id=grub)
                venue = veg.n_venue
            else:
                veg = Both.objects.get(gm_id=grub)
                venue = veg.non_veg_venue
            abcd = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up", meal="Non Veg")
            k = len(abcd) // 99
            for q in range(k + 1):
                a = []
                students = abcd[q * 99:(q + 1) * 99]
                for j in students:
                    if (j.mail != "Sent2"):
                        a.append(str(j.user_id) + "@pilani.bits-pilani.ac.in")
                        j.mail = "Sent2"
                        j.save()
                if (len(a) > 0):
                    subject, from_email = str(grub.name) + " | " + e + \
                                          " | Non Veg | " + venue, 'ssms.pilani@gmail.com'
                    text_content = 'This is an important message.'
                    html_content = "<body><p>This is to remind you that you that you have been signed up for <strong> " + \
                                   str(
                                       grub.name) + "</strong> that is going to take place on <strong>" + e + "</strong> at the " + venue + ".\
					You are required \
					to collect your stubs from the PitStop counter in your mess during meal timings today.</p>\
	<strong><p>Entry into the grub shall not be allowed if you are not wearing the wristband.</p></strong>\
					<p>" + getspotsigning + " </p><p>Thank you.</p>\
					<p>Regards,</p>\
					<p>Grub Committee, SSMS</p></body>"
                    # print a
                    msg = EmailMultiAlternatives(subject, text_content, from_email, cc=a, bcc=[
                        "f2015040@pilani.bits-pilani.ac.in"])
                    msg.attach_alternative(html_content, "text/html")
                    try:
                        msg.send(fail_silently=False)
                        count = count + len(a)
                        datemail.mails = datemail.mails + len(a)
                        datemail.save()
                    except Exception as e:
                        for j in students:
                            j.mail = "Sent"
                            j.save()
                        left = len(allstu) - count
                        data = "Only " + str(count) + " -Non Veg- mails were sent succesfully. " + \
                               str(left) + " mails are left to be send. Error " + str(e)
                        return data
        grub.mails = "Sent2"
        grub.save()
        data = "All the mails (" + str(len(allstu)) + ") were sent succesfully"
        return data
    except Exception as e:
        print("here")
        data = "Internal Error " + str(e)
        return data


def ssms_grub_sendmail2(request, gmid):
    grubid = request.GET.get('grubid')
    datemail = DateMailStatus.objects.get(date=datetime.now())
    try:
        grub = Grub.objects.get(name=str(grubid))
        print(gmid)
        if (grub.meal == "Veg"):
            forloop = 1
        elif (grub.meal == "Non Veg"):
            forloop = 2
        else:
            forloop = 3  # both  1 and two
        d = datetime.strptime(str(grub.date), '%Y-%m-%d')
        e = date.strftime(d, "%d %B %Y")
        count = 0
        if (grub.spot_signing == "Yes"):
            getspotsigning = "Limited on spot signings will be available. Please carry your ID cards for the same."
        else:
            getspotsigning = ""
        allstu = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up")
        allgrubbatch = Batch.objects.filter(gm_id=grub)
        if (len(allgrubbatch) == 0):
            print("here2")
            # return JsonResponse({"is_taken" : "here"})
            # {'is_taken': "No batch allocated. Please allocate batch first." }
            data = ssms_grub_sendmail3(grub, forloop, datemail, d, e, getspotsigning)
            return JsonResponse({"is_taken": data})
        if (forloop == 1 or forloop == 3):
            if (forloop == 1):
                veg = Veg.objects.get(gm_id=grub)
                venue = veg.v_venue
            else:
                veg = Both.objects.get(gm_id=grub)
                venue = veg.veg_venue
            all_batch = Batch.objects.filter(gm_id=grub, meal="Veg")
            print(all_batch)
            for i in all_batch:
                abcd = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up", batch=i)
                k = len(abcd) // 99
                for q in range(k + 1):
                    a = []
                    students = abcd[q * 99:(q + 1) * 99]
                    for j in students:
                        if (j.mail != "Sent2"):
                            a.append(str(j.user_id) + "@pilani.bits-pilani.ac.in")
                            j.mail = "Sent2"
                            j.save()
                    if (len(a) > 0):
                        subject, from_email = str(
                            grub.name) + " | " + e + " | Veg | " + "Batch " + str(i.batch_name), 'ssms.pilani@gmail.com'
                        text_content = 'This is an important message.'
                        html_content = "<body><p>This is to remind you that you that you have been signed up for <strong> " + \
                                       str(grub.name) + "</strong> that is going to take place on <strong>" + e + "</strong>. You are required \
						to collect your stubs from the PitStop counter in your mess during meal timings today.</p>\
						<p>Also, please note the following details.</p>\
						<p><strong>Batch: </strong>" + i.batch_name + "</p>\
						<p><strong>Wristband: </strong>" + i.color + "</p>\
						<p><strong>Timings: </strong>" + i.timing + "</p>\
						<p><strong>Venue: </strong>" + venue + "</p>\
		<strong><p>Entry into the grub shall not be allowed if you are not wearing the wristband.</p></strong>\
						<p>" + getspotsigning + " </p><p>Thank you.</p>\
						<p>Regards,</p>\
						<p>Grub Committee, SSMS</p></body>"
                        print(a)
                        msg = EmailMultiAlternatives(subject, text_content, from_email, cc=a, bcc=[
                            "f2015040@pilani.bits-pilani.ac.in"])
                        msg.attach_alternative(html_content, "text/html")
                        try:
                            msg.send(fail_silently=False)
                            count = count + len(a)
                            datemail.mails = datemail.mails + len(a)
                            datemail.save()
                        except Exception as e:
                            for j in students:
                                j.mail = "Sent"
                                j.save()
                            left = len(allstu) - count
                            data = {'is_taken': "Only " + str(count) + " mails were sent succesfully. " + str(
                                left) + " mails are left to be send. Error " + str(e)}
                            return JsonResponse(data)
        if (forloop == 2 or forloop == 3):
            if (forloop == 1):
                veg = NonVeg.objects.get(gm_id=grub)
                venue = veg.n_venue
            else:
                veg = Both.objects.get(gm_id=grub)
                venue = veg.non_veg_venue
            all_batch = Batch.objects.filter(gm_id=grub, meal="Non Veg")
            print(all_batch)
            for i in all_batch:
                abcd = Grub_Student.objects.filter(gm_id=grub.gm_id, status="Signed Up", batch=i)
                k = len(abcd) // 99
                for q in range(k + 1):
                    a = []
                    students = abcd[q * 99:(q + 1) * 99]
                    for j in students:
                        if (j.mail != "Sent2"):
                            a.append(str(j.user_id) + "@pilani.bits-pilani.ac.in")
                            j.mail = "Sent2"
                            j.save()
                    if (len(a) > 0):
                        subject, from_email = str(
                            grub.name) + " | " + e + " | Non Veg | " + "Batch " + str(
                            i.batch_name), 'ssms.pilani@gmail.com'
                        text_content = 'This is an important message.'
                        html_content = "<body><p>This is to remind you that you that you have been signed up for <strong> " + \
                                       str(grub.name) + "</strong> that is going to take place on <strong>" + e + "</strong>. You are required \
						to collect your stubs from the PitStop counter in your mess during meal timings today.</p>\
						<p>Also, please note the following details.</p>\
						<p><strong>Batch: </strong>" + i.batch_name + "</p>\
						<p><strong>Wristband: </strong>" + i.color + "</p>\
						<p><strong>Timings: </strong>" + i.timing + "</p>\
						<p><strong>Venue: </strong>" + venue + "</p>\
		<strong><p>Entry into the grub shall not be allowed if you are not wearing the wristband.</p></strong>\
						<p>" + getspotsigning + " </p><p>Thank you.</p>\
						<p>Regards,</p>\
						<p>Grub Committee, SSMS</p></body>"
                        print(a)
                        msg = EmailMultiAlternatives(subject, text_content, from_email, cc=a, bcc=[
                            "f2015040@pilani.bits-pilani.ac.in"])
                        msg.attach_alternative(html_content, "text/html")
                        try:
                            msg.send(fail_silently=False)
                            count = count + len(a)
                            datemail.mails = datemail.mails + len(a)
                            datemail.save()
                        except Exception as e:
                            for j in students:
                                j.mail = "Sent"
                                j.save()
                            left = len(allstu) - count
                            data = {'is_taken': "Only " + str(count) + " mails were sent succesfully. " + str(
                                left) + " mails are left to be send. Error " + str(e)}
                            return JsonResponse(data)
        grub.mails = "Sent2"
        grub.save()
        data = {'is_taken': "All the mails (" + str(len(abcd)) + ") were sent succesfully"}
        return JsonResponse(data)
    except Exception as e:
        print("here")
        data = {'is_taken': "Internal Error " + str(e)}
        return JsonResponse(data)


def ssms_grub_sendmail(request, gmid):
    if request.user.is_superuser:
        context_dict = {}
        try:
            grub = Grub.objects.get(gm_id=gmid)
            context_dict["grub"] = grub
            stud = Grub_Student.objects.filter(gm_id=gmid, status="Signed Up")
            registered = len(stud)
            context_dict["registered"] = registered
            if (DateMailStatus.objects.filter(date=datetime.now()).exists()):
                datemail = DateMailStatus.objects.get(date=datetime.now())
                context_dict["datemail"] = datemail
            else:
                datemail = DateMailStatus.objects.create(date=datetime.now(), mails=0)
                context_dict["datemail"] = datemail
            mailsleft = 1000 - datemail.mails
            context_dict["mailsleft"] = mailsleft
        except Exception as e:
            print(e)
            context_dict["error"] = e
        return render(request, 'ssms/ssms_grub_sendmail.html', context_dict)
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


def index(request):
    if not request.user.is_authenticated():
        return render(request, 'ssms/index.html')
    else:

        try:
            coord = Grub_Coord.objects.get(user=request.user)
            grub = Grub.objects.filter(cg_id=coord.cg_id).order_by('-date')
            context_dict = {'grub': grub, 'coord': coord}
            return render(request, 'ssms/index.html', context_dict)
        except Grub_Coord.DoesNotExist:
            context_dict = {}
            pass
        try:
            stud = Student.objects.get(user_id=str(request.user))
            context_dict = {'student': stud}
            return render(request, 'ssms/index.html', context_dict)
        except Student.DoesNotExist:
            context_dict = {}
            pass
        return render(request, 'ssms/index.html', context_dict)


def ssms_login(request):  # admin login
    context_dict = {}
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)

        if user and user.is_superuser:
            if user.is_active:
                login(request, user)
                return HttpResponseRedirect('/ssms/')
            else:
                return HttpResponse("Your ssms account is disabled.")
        else:

            context_dict['invalid'] = "Invalid login details supplied."
            return render(request, 'ssms/ssms_login.html', context_dict)

    else:
        return render(request, 'ssms/ssms_login.html')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def ssms_register(request):
    if request.user.is_superuser:
        registered = False
        if request.method == 'POST':
            user_form = Grub_CoordUserForm(data=request.POST)
            profile_form = Grub_CoordUserProfileForm(data=request.POST)
            if user_form.is_valid() and profile_form.is_valid():
                user = user_form.save(commit=False)
                user.set_password(user.password)
                user.is_staff = True

                profile = profile_form.save(commit=False)

                profile.status = "Active"
                profile.date = datetime.now()
                profile.reg_by = str(request.user)
                user.save()
                profile.user = user
                profile.save()
                registered = True

            else:
                print(user_form.errors)
                print(profile_form.errors)

        else:
            user_form = Grub_CoordUserForm()
            profile_form = Grub_CoordUserProfileForm()
        return render(request, 'ssms/ssms_register.html',
                      {'user_form': user_form, 'profile_form': profile_form, 'registered': registered})
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def ssms_grubinfo(request, gmid):  # grub info accessible to admin
    if request.user.is_superuser:
        date_diff = datechecker(gmid)
        context_dict = {}
        try:
            grub = Grub.objects.get(gm_id=gmid)
            if grub.meal == 'Veg':
                grub_meal = Veg.objects.get(gm_id=gmid)
                grub_type = 0
            elif grub.meal == 'Non Veg':
                grub_meal = NonVeg.objects.get(gm_id=gmid)
                grub_type = 1
            elif grub.meal == 'Both':
                grub_meal = Both.objects.get(gm_id=gmid)
                grub_type = 2
            context_dict['grub'] = grub
            context_dict['meal'] = grub_meal
            context_dict['grub_type'] = grub_type
            context_dict['date_diff'] = date_diff
        except Grub.DoesNotExist:
            pass
        return render(request, 'ssms/ssms_grubinfo.html', context_dict)
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


def ssms_grubeditdeadline(request, gmid):  # allows admin to edit deadline of grub
    if request.user.is_superuser:
        if request.method == 'POST':
            grub = Grub.objects.get(gm_id=gmid)
            if request.POST["deadline"] != "":
                grub.deadline = request.POST["deadline"]
            if request.POST["deadline2"] != "":
                grub.deadline2 = request.POST["deadline2"]
            grub.save()
            return HttpResponseRedirect('/ssms/ssms/grub/' + gmid)
        else:
            grub = Grub.objects.get(gm_id=gmid)
        return render(request, 'ssms/ssms_grubeditdeadline.html', {'grub': grub})
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def ssms_grub_list(request):
    if request.user.is_superuser:
        grub_list = Grub.objects.order_by('-date')[:]
        coord_list = Grub_Coord.objects.all()
        context_dict = {"grub": grub_list, "coord": coord_list}
        return render(request, 'ssms/ssms_grublist.html', context_dict)
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def ssms_student_table(request, gmid):
    try:
        grub = Grub.objects.get(gm_id=gmid)

        if request.user.is_staff:
            coord = Grub_Coord.objects.get(cg_id=grub.cg_id.cg_id)
            if not request.user == coord.user:
                return HttpResponseRedirect("/ssms")

        elif request.user.is_superuser:
            pass

        else:
            return HttpResponseRedirect("/ssms")
        # in case a grub has both meals available -- only then will the number of veg
        # and non veg students be displayed
        # in case the grub has veg or non-veg status then only the registered number
        # of students is displayed
        grub = Grub.objects.get(gm_id=gmid)
        stud = Grub_Student.objects.filter(gm_id=gmid)
        registered = len(stud.filter(status="Signed Up"))
        out = len(stud.filter(status="Opted Out"))
        vegreg = nonvegreg = ""
        if grub.meal == "Both":
            vegreg = len(stud.filter(status="Signed Up", meal="Veg"))
            nonvegreg = len(stud.filter(status="Signed Up", meal="Non Veg"))
        context_dict = {"stud": stud, "reg": registered, "out": out,
                        "gmid": gmid, "grub": grub, 'vegreg': vegreg, 'nonvegreg': nonvegreg}
        return render(request, 'ssms/dynamic_table.html', context_dict)
    except Grub.DoesNotExist:
        pass
    return HttpResponseRedirect("/ssms")


def ssms_coord_active(request, cgid):
    if request.user.is_superuser:
        a = Grub_Coord.objects.get(cg_id=cgid)
        b = User.objects.get(username=a.user)
        c = Grub.objects.filter(cg_id=cgid)
        for i in c:
            i.status = "Active"
            i.save()
        b.is_staff = True
        b.is_active = True
        a.status = "Active"
        a.save()
        b.save()
        return HttpResponseRedirect("/ssms/ssms/grublist/")
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


def ssms_coord_inactive(request, cgid):
    if request.user.is_superuser:
        a = Grub_Coord.objects.get(cg_id=cgid)
        b = User.objects.get(username=a.user)
        c = Grub.objects.filter(cg_id=cgid)
        for i in c:
            i.status = "Inactive"
            i.save()
        b.is_staff = False
        b.is_active = False
        a.status = "Inactive"
        a.save()
        b.save()
        return HttpResponseRedirect("/ssms/ssms/grublist/")
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


def ssms_grub_spot_signing(request, gmid):  # toggle grub spot signing setting
    if request.user.is_superuser:
        grub = Grub.objects.get(gm_id=gmid)
        if grub.spot_signing == "Yes":
            grub.spot_signing = "No"
        else:
            grub.spot_signing = "Yes"
        grub.save()
        return HttpResponseRedirect('/ssms/ssms/grub/' + gmid)
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


def ssms_student_cancel(request, gmid):
    context_dict = {}
    if request.user.is_superuser:  # and not request.user.is_superuser - Allow superuser to access
        try:
            grub = Grub.objects.get(gm_id=gmid)
            context_dict["grub"] = grub
            e = datechecker(gmid)
            context_dict["e"] = e
            done = 0
            context_dict["done"] = done
            if e == 1:
                if request.method == 'POST':
                    try:
                        grubstu = Grub_Student.objects.get(
                            student_id=request.POST['student_id'], gm_id=grub)
                        if grubstu.status == "Opted Out":
                            context_dict["invalid"] = "Student already opted out."
                            return render(request, 'ssms/ssms_student_cancel.html', context_dict)
                        grubstu.status = "Opted Out"
                        grubstu.save()
                        done = 1
                        context_dict["done"] = done
                        return render(request, 'ssms/ssms_student_cancel.html', context_dict)
                    except:
                        context_dict["invalid"] = "Student is not registered in the grub."
                        return render(request, 'ssms/ssms_student_cancel.html', context_dict)
                else:
                    return render(request, 'ssms/ssms_student_cancel.html', context_dict)
            else:
                return render(request, 'ssms/ssms_student_cancel.html', context_dict)
        except Grub.DoesNotExist:
            pass
            return HttpResponseRedirect("/ssms")
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


def export_data(request, gmid):
    if request.user.is_staff:
        if request.method == "POST":
            try:
                bhawan = request.POST.get('bhawan')
                if bhawan == "All":
                    grub_students = Grub_Student.objects.filter(gm_id=gmid, status="Signed Up")
                else:
                    grub_students = Grub_Student.objects.filter(gm_id=gmid, status="Signed Up", bhawan=bhawan)

                grub = Grub.objects.get(gm_id=gmid)
                data = [
                    [
                        student.user_id,
                        student.name,
                        student.student_id,
                        student.meal,
                        student.bhawan,
                        student.room,
                        student.batch.batch_name if student.batch else ""
                    ]
                    for student in grub_students
                ]

                workbook = xlsxwriter.Workbook('media/' + grub.name + '_' + bhawan + '_grublist.xlsx')
                worksheet = workbook.add_worksheet()
                worksheet.set_column('A:A', 15)
                worksheet.set_column('B:B', 25)
                worksheet.set_column('C:C', 20)
                worksheet.set_column('D:D', 15)
                worksheet.set_column('E:E', 15)
                worksheet.set_column('F:F', 15)
                worksheet.set_column('G:G', 5)
                bold = workbook.add_format({'bold': 1})
                worksheet.write('A1', 'User ID', bold)
                worksheet.write('B1', 'Name', bold)
                worksheet.write('C1', 'BITS ID', bold)
                worksheet.write('D1', 'Meal Type', bold)
                worksheet.write('E1', 'Bhawan', bold)
                worksheet.write('F1', 'Room No.', bold)
                worksheet.write('G1', 'Batch', bold)
                row = 1
                col = 0
                for i in data:
                    worksheet.write_string(row, col, i[0])
                    worksheet.write_string(row, col + 1, i[1])
                    worksheet.write_string(row, col + 2, i[2])
                    worksheet.write_string(row, col + 3, i[3])
                    worksheet.write_string(row, col + 4, i[4])
                    worksheet.write_string(row, col + 5, i[5])
                    worksheet.write_string(row, col + 6, i[6])
                    row += 1
                workbook.close()
                return HttpResponseRedirect('media/' + grub.name + '_' + bhawan + '_grublist.xlsx')
            except Grub.DoesNotExist:
                pass
                return HttpResponseRedirect("/ssms")
        else:
            return HttpResponseRedirect("/ssms")

    else:
        return HttpResponseRedirect('/ssms/')


def coord_login(request):
    context_dict = {}
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)

        if user and user.is_staff and not user.is_superuser:
            if user.is_active:  # add status  choice here
                login(request, user)
                return HttpResponseRedirect('/ssms/')
            else:
                return HttpResponse("Your ssms account is disabled.")
        else:
            context_dict['invalid'] = "Invalid login details supplied."
            print("Invalid login details: {0}, {1}".format(username, password))
            return render(request, 'ssms/coord_login.html', context_dict)

    else:
        return render(request, 'ssms/coord_login.html', {})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@user_passes_test(lambda user: user.is_staff or user.is_superuser, login_url='/ssms/coord/login/')
def coord_grub_register(request):
    """registers a new grub or returns errors in the form"""
    if request.method == 'POST':
        grub_form = GrubForm(request.POST, request.FILES)
        veg_form = VegForm(request.POST, request.FILES)
        non_veg_form = NonVegForm(request.POST, request.FILES)
        both_form = BothForm(request.POST, request.FILES)
        done = False

        if grub_form.is_valid():
            grub = grub_form.save(commit=False)
            grub.meal = request.POST.get('mealtype')
            grub.date = request.POST.get('grubdate')
            grub.reg_date = datetime.now()
            grub.cg_id = Grub_Coord.objects.get(user=request.user)
            grub.save()
            grub.deadline = grub.date - timedelta(days=2)
            grub.deadline2 = grub.date - timedelta(days=4)
            grub.save()

            if veg_form.is_valid():
                photo = veg_form.save(commit=False)
                photo.name = request.POST.get('name')
                photo.veg_price = request.POST.get('v_price')
                photo.veg_venue = request.POST.get('v_venue')
                grub_coord = Grub.objects.get(gm_id=grub.gm_id)
                photo.gm_id = grub_coord
                if 'v_images' in request.FILES:
                    photo.v_images = request.FILES['v_images']
                    photo.save()
                    done = True
            elif non_veg_form.is_valid():
                photo = non_veg_form.save(commit=False)
                photo.veg_venue = request.POST.get('n_venue')
                photo.non_veg_venue = request.POST.get('n_venue')
                grub_coord = Grub.objects.get(gm_id=grub.gm_id)
                photo.gm_id = grub_coord
                if 'n_images' in request.FILES:
                    photo.n_images = request.FILES['n_images']
                    photo.save()
                    done = True
            elif both_form.is_valid():
                photo = both_form.save(commit=False)
                photo.veg_price = request.POST.get('veg_price')
                photo.non_veg_price = request.POST.get('non_veg_price')
                photo.veg_venue = request.POST.get('veg_venue')
                photo.non_veg_venue = request.POST.get('non_veg_venue')
                grub_coord = Grub.objects.get(gm_id=grub.gm_id)
                photo.gm_id = grub_coord
                if 'veg_images' in request.FILES and 'non_veg_images' in request.FILES:
                    photo.veg_images = request.FILES['veg_images']
                    photo.non_veg_images = request.FILES['non_veg_images']
                    photo.save()
                    done = True

        context = {
            'form': grub_form,
            'form1': veg_form,
            'form2': non_veg_form,
            'form3': both_form,
            'done': done,
        }

    else:
        context = {
            'form': GrubForm(),
            'form1': VegForm(),
            'form2': NonVegForm(),
            'form3': BothForm(),
            'done': False
        }

    return render(request, 'ssms/coord_grub_register.html', context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@user_passes_test(lambda user: user.is_superuser or user.is_staff, login_url='ssms/coord/login')
def coord_upload(request, gmid):

    """
    this function is used to validate the file input and upload it with details for the grub students
    :param request:
    :param gmid:
    Validation process:
    authenticate the user - grub coordinator or admin only
    check for id if the user is a staff
    checks for deadlines and provisions for on spot signings
    checks if the method request is a POST and checks if a file has been passed
    check if an excel file has been passed
    checks if the file has a valid spreadsheet extension
    check if the form is valid
    if any of the above cases is not satisfied redirect and show an appropriate notice
    :return:
    """
    try:
        grub = Grub.objects.get(gm_id=gmid)
        form = ExcelUpload(instance=grub)
        grub_date = datechecker(gmid)
        if request.user.is_staff:
            coord = Grub_Coord.objects.get(cg_id=grub.cg_id.cg_id)
            if not request.user == coord.user:
                return HttpResponseRedirect("/ssms")
        else:
            return HttpResponseRedirect('/ssms/coord/login/')

        if not (grub_date == 2 or (grub_date == 4 and grub.spot_signing == "Yes")):
            return render(request, 'ssms/coord_upload.html', {"e": grub_date, 'grub': grub})
        elif not (request.method == 'POST' and request.FILES):
            form = ExcelUpload(instance=grub)
        elif not ('excel' in request.FILES and (str(request.FILES['excel'].name).endswith(('.xls', '.xlsx', '.csv'))) and
                                                ExcelUpload(request.POST, request.FILES, instance=grub).is_valid()):

            context = {'form': form, 'grub': grub, "e": grub_date, "invalid": "No Spreadsheet Detected"}
            return render(request, 'ssms/coord_upload.html', context)

        Grub_Invalid_Students.objects.filter(gm_id=gmid).delete()
        d = Grub.objects.filter(gm_id=gmid)[0]
        form = ExcelUpload(request.POST, request.FILES, instance=grub)
        photo = form.save(commit=False)

        def choice_func(row):
            row[0] = row[0].upper()
            a = row[0]
            # Change to 13 if "P" is too be included
            a = str(a)[:12]
            try:
                student = Student.objects.get(bits_id=a)
                row.append(student.user_id)
                row.append(student.name)
                row.append(student.bhawan)
                row.append(student.room_no)

            except:
                return None
            if str(row[1]).lower() == "veg":
                row[1] = "Veg"
            elif str(row[1]).lower() == "non veg":
                row[1] = "Non Veg"
            row.append("Signed Up")
            row.append(d)
            return row

        files = request.FILES['excel']
        files.save_to_database(
            model=Grub_Student,
            initializer=choice_func,
            mapdict=['student_id', 'meal', 'user_id',
                     'name', 'bhawan', 'room', 'status', 'gm_id']
        )
        photo.excel = files
        photo.save()
        if len(Grub_Invalid_Students.objects.filter(gm_id=grub)) > 0:
            return HttpResponseRedirect("/ssms/invalid_ids/" + gmid)
        else:
            return HttpResponseRedirect("/ssms/stats/" + gmid)

    except Grub.DoesNotExist:
        return HttpResponseRedirect("/ssms")

    except Exception as exception:
        invalid = "The following error occured : \n" + str(exception)
        return render(request, 'ssms/coord_upload.html',
                      {'form': form, 'grub': grub, "e": grub_date, "invalid": invalid})

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@user_passes_test(lambda user: user.is_superuser, login_url='/ssms/ssms/login/')
def ssms_mem_upload(request, gmid):
    try:
        grub = Grub.objects.get(gm_id=gmid)
        form = ExcelUpload(instance=grub)
        deadline_status = datechecker(gmid)
        if deadline_status != 4:  # if grub is not over return
            return render(request, 'ssms/coord_mem_upload.html', {"e": deadline_status, 'grub': grub})

        if not (request.method == 'POST' and request.FILES):  # pass only if request is post and file is uploaded
            form = ExcelUpload(instance=grub)
            return render(request, 'ssms/coord_mem_upload.html', {'form': form, 'grub': grub, "e": deadline_status})

        if 'excel' not in request.FILES:
            invalid = "No file uploaded."
            return render(request, 'ssms/coord_mem_upload.html',
                          {'form': form, 'grub': grub, "e": deadline_status, "invalid": invalid})

        filename = str(request.FILES['excel'].name)
        if not filename.endswith(('.xls', '.xlsx', '.csv')):
            invalid = "Unsupported File type."
            return render(request, 'ssms/coord_mem_upload.html',
                          {'form': form, 'grub': grub, "e": deadline_status, "invalid": invalid})

        grub = Grub.objects.filter(gm_id=gmid)[0]
        form = ExcelUpload(request.POST, request.FILES, instance=grub)
        if not form.is_valid():
            invalid = "Invalid File type."
            return render(request, 'ssms/coord_mem_upload.html',
                          {'form': form, 'grub': grub, "e": deadline_status, "invalid": invalid})

        photo = form.save(commit=False)

        def choice_func(row):
            # Change to 13 if "P" is too be included
            a = row[0].upper()[:12]
            try:
                stu = Grub_Student.objects.get(
                    student_id=str(a), gm_id=gmid)
                stu.status = "Member"
                stu.save()
                row.append(grub)
                return row
            except Exception as e:
                try:
                    stu = Grub_Student.objects.get(
                        student_id=str(a) + "P", gm_id=gmid)
                    stu.status = "Member"
                    stu.save()
                    row.append(grub)
                    return row
                except Exception as e:
                    return row

        files = request.FILES['excel']
        files.save_to_database(
            model=Grub_Member,
            initializer=choice_func,
            mapdict=['student_id', 'meal', 'gm_id']
        )
        photo.excel = files
        photo.save()

    except Grub.DoesNotExist:
        return HttpResponseRedirect("/ssms")

    except Exception as e:
        invalid = "The following error occured : \n" + str(e)
        return render(request, 'ssms/coord_mem_upload.html', {'form': form, 'grub': grub, "e": deadline_status, "invalid": invalid})


@user_passes_test(lambda user: user.is_staff or user.is_superuser, login_url='/ssms/coord/login/')
def coord_invalid_ids(request, gmid):
    try:
        # grub coord can only view stats of their registered grub
        if request.user.is_staff:
            grub = Grub.objects.get(gm_id=gmid)
            coord = Grub_Coord.objects.get(cg_id=grub.cg_id.cg_id)
            if request.user != coord.user:
                return HttpResponseRedirect("/ssms")

        invalid_stud = Grub_Invalid_Students.objects.filter(gm_id=gmid)
        context_dict = {"stud": invalid_stud, "invalidno": len(invalid_stud), "gmid": gmid, "grub": grub}
        return render(request, 'ssms/invalid_student_table.html', context_dict)

    except Grub.DoesNotExist:
        return HttpResponseRedirect("/ssms")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coord_student_register(request, gmid):
    if request.user.is_staff:  # and not request.user.is_superuser - Allow superuser to access
        try:
            grub = Grub.objects.get(gm_id=gmid)
            form = ExcelUpload(instance=grub)
            e = datechecker(gmid)
            coord = Grub_Coord.objects.get(cg_id=grub.cg_id.cg_id)
            if request.user == coord.user or request.user.is_superuser:
                done = 0
                if e == 2:
                    if request.method == 'POST':
                        form = CoordStudentRegForm(request.POST)
                        if form.is_valid():
                            photo = form.save(commit=False)
                            photo.student_id = photo.student_id.upper()
                            try:
                                gs = Grub_Student.objects.get(
                                    student_id=photo.student_id, gm_id=gmid)
                                invalid = "Student already registered."
                                return render(request, 'ssms/coord_student_register.html',
                                              {'form': form, 'done': done, 'grub': grub, 'e': e, "invalid": invalid})
                            except Grub_Student.DoesNotExist:
                                pass
                            photo.status = "Signed Up"
                            photo.meal = request.POST.get('mealtype')
                            a = photo.student_id
                            if (a[4].upper() == "H" or a[4].upper() == "P"):
                                if (int(a[0:4]) < 2017):
                                    photo.user_id = a[4] + a[0:4] + a[9:12]
                                else:
                                    photo.user_id = a[4] + a[0:4] + a[8:12]
                            else:
                                if (int(a[0:4]) < 2017):
                                    photo.user_id = 'f' + a[0:4] + a[9:12]
                                else:
                                    photo.user_id = 'f' + a[0:4] + a[8:12]
                            photo.user_id = photo.user_id.lower()
                            print(photo.user_id)
                            try:
                                d = Student.objects.get(bits_id=a)
                                photo.room = d.room_no
                                photo.bhawan = d.bhawan
                                photo.name = d.name
                                b = Grub.objects.filter(gm_id=gmid)[0]
                                photo.gm_id = b
                                photo.save()
                                done = 1
                            except Student.DoesNotExist:
                                pass
                                invalid = "Invalid ID"
                                return render(request, 'ssms/coord_student_register.html',
                                              {'form': form, 'done': done, 'grub': grub, 'e': e, "invalid": invalid})

                        else:
                            print(form.errors)
                    else:
                        form = CoordStudentRegForm()
                        grub = Grub.objects.get(gm_id=gmid)

                    return render(request, 'ssms/coord_student_register.html',
                                  {'form': form, 'done': done, 'grub': grub, 'e': e})
                else:
                    return render(request, 'ssms/coord_student_register.html', {'grub': grub, 'e': e})
            else:
                return HttpResponseRedirect("/ssms")
        except Grub.DoesNotExist:
            pass
            return HttpResponseRedirect("/ssms")
    else:
        return HttpResponseRedirect('/ssms/coord/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@user_passes_test(lambda user: user.is_superuser or user.is_staff, login_url='ssms/coord/login')
def coord_view_grub(request, gmid):
    try:
        grub = Grub.objects.get(gm_id=gmid)
        grub_date = datechecker(gmid)
        if request.user.is_staff :
            coord = Grub_Coord.objects.get(cg_id=grub.cg_id.cg_id)
            if not request.user == coord.user:
                return HttpResponseRedirect("/ssms")

        context_dict = {}
        try:
            grub = Grub.objects.get(gm_id=gmid)
            if grub.meal == 'Veg':
                meal = Veg.objects.get(gm_id=gmid)
                meal_type = 0
            elif grub.meal == 'Non Veg':
                meal = NonVeg.objects.get(gm_id=gmid)
                meal_type = 1
            elif grub.meal == 'Both':
                meal = Both.objects.get(gm_id=gmid)
                meal_type = 2
            context_dict['grub'] = grub
            context_dict['meal'] = meal
            context_dict['i'] = meal_type
            context_dict['e'] = grub_date
        except Grub.DoesNotExist:
            pass
        return render(request, 'ssms/coord_grubinfo.html', context_dict)

    except Grub.DoesNotExist:
        pass
    return HttpResponseRedirect("/ssms")

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coord_grub_edit(request, gmid):
    if request.user.is_staff and not request.user.is_superuser:
        try:
            grub = Grub.objects.get(gm_id=gmid)
            form = ExcelUpload(instance=grub)
            e = datechecker(gmid)
            coord = Grub_Coord.objects.get(cg_id=grub.cg_id.cg_id)
            if request.user == coord.user:
                done = 0
                inst = Grub.objects.get(gm_id=gmid)
                grub = Grub.objects.get(gm_id=gmid)
                if request.method == 'POST':
                    form = GrubFormEdit(request.POST, instance=inst)
                    if form.is_valid():
                        grub = form.save(commit=False)
                        grub.save()
                        done = 1
                    else:
                        print(form.errors)
                else:
                    form = GrubFormEdit(instance=inst)
                return render(request, 'ssms/coord_grub_edit.html', {'form': form, 'done': done, 'grub': grub})
            else:
                return HttpResponseRedirect("/ssms")
        except Grub.DoesNotExist:
            pass
            return HttpResponseRedirect("/ssms")
    else:
        return HttpResponseRedirect('/ssms/coord/login/')


# Coord do not have permission to cancel/activate grubs


def ssms_grub_inactive(request, gmid):
    if request.user.is_superuser:
        a = Grub.objects.get(gm_id=gmid)
        a.status = "Inactive"
        a.save()
        return HttpResponseRedirect("/ssms/ssms/grub/" + gmid)
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


def ssms_grub_active(request, gmid):
    if request.user.is_superuser:
        a = Grub.objects.get(gm_id=gmid)
        a.status = "Active"
        a.save()
        return HttpResponseRedirect("/ssms/ssms/grub/" + gmid)
    else:
        return HttpResponseRedirect('/ssms/ssms/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def student_upcoming_grubs(request):
    if request.user.is_authenticated() and not request.user.is_staff:
        grub_list = Grub.objects.filter(status="Active").order_by('-date')[:]
        context_dict = {"grub": grub_list}
        return render(request, 'ssms/student_grublist.html', context_dict)
    else:
        return HttpResponseRedirect("/ssms/")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def student_grub_register(request, gmid):
    context_dict = {'gmid': gmid}
    if request.user.is_authenticated() and not request.user.is_staff:
        try:
            grub = Grub.objects.get(gm_id=gmid, status="Active")
            grubcoord = Grub_Coord.objects.get(cg_id=grub.cg_id.cg_id)
            if grub.meal == 'Veg':
                grub_meal = Veg.objects.get(gm_id=gmid)
                meal_type = 0
            elif grub.meal == 'Non Veg':
                grub_meal = NonVeg.objects.get(gm_id=gmid)
                meal_type = 1
            elif grub.meal == 'Both':
                grub_meal = Both.objects.get(gm_id=gmid)
                meal_type = 2
            try:
                a = Grub_Student.objects.get(gm_id=gmid, user_id=str(request.user))
                context_dict['student'] = a
            except Grub_Student.DoesNotExist:
                pass
            date_status = datechecker(gmid)
            context_dict['grub'] = grub
            context_dict['grubcoord'] = grubcoord
            context_dict['meal'] = grub_meal
            context_dict['i'] = meal_type
            context_dict['e'] = date_status
        except Grub.DoesNotExist:
            pass
        return render(request, 'ssms/student_grubinfo.html', context_dict)
    else:
        return render(request, 'ssms/student_grubinfo.html', context_dict)
        return HttpResponseRedirect("/soc/login/google-oauth2/?next=/ssms/student/grub/" + gmid)


def student_grub_register2(request, gmid):  # register for veg
    if not (request.user.is_authenticated() and not request.user.is_staff):
        return HttpResponseRedirect("/ssms/")
    try:
        grub = Grub.objects.get(gm_id=gmid, status="Active")
        diff = datechecker(gmid)
        if not (diff == 1 or diff == 2):  # register student for grub only within deadline
            return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")

        student = Student.objects.get(user_id=str(request.user))
        try:  # if exists change Veg to signed up
            grub_student = Grub_Student.objects.get(gm_id=gmid, user_id=str(request.user))
            grub_student.meal = "Veg"
            grub_student.status = "Signed Up"
            grub_student.save()
        except Grub_Student.DoesNotExist:  # if student is not registered create new object and register
            Grub_Student.objects.create(gm_id=grub, user_id=str(request.user), student_id=str(student.bits_id),
                                        meal="Veg", status="Signed Up",
                                        room=student.room_no, bhawan=student.bhawan,
                                        name=student.name)  # if +"P" here in str(d.bits_id)+"P" change the whole thing accordingly
        return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")

    except Grub.DoesNotExist:
        return HttpResponseRedirect("/ssms/")


def student_grub_register3(request, gmid):  # register for non veg
    if request.user.is_authenticated() and not request.user.is_staff:
        try:
            grub = Grub.objects.get(gm_id=gmid, status="Active")
            d = datechecker(gmid)
            if (d == 1 or d == 2):
                a = Grub.objects.filter(gm_id=gmid)[0]
                d = Student.objects.get(user_id=str(request.user))
                try:
                    b = Grub_Student.objects.get(gm_id=gmid, user_id=str(request.user))
                    b.meal = "Non Veg"
                    b.status = "Signed Up"
                    b.save()
                except Grub_Student.DoesNotExist:
                    Grub_Student.objects.create(gm_id=a, user_id=str(request.user), student_id=str(d.bits_id),
                                                meal="Non Veg", status="Signed Up",
                                                room=d.room_no, bhawan=d.bhawan,
                                                name=d.name)  # if +"P" here in str(d.bits_id)+"P" change the whole thing accordingly

                return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")
            else:
                return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")
        except Grub.DoesNotExist:
            pass
            return HttpResponseRedirect("/ssms/")
    else:
        return HttpResponseRedirect("/ssms/")


def student_grub_cancel(request, gmid):
    if request.user.is_authenticated() and not request.user.is_staff:
        try:
            grub = Grub.objects.get(gm_id=gmid, status="Active")
            d = datechecker(gmid)
            if d == 1:
                try:
                    a = Grub_Student.objects.get(gm_id=gmid, user_id=str(request.user))
                    a.status = "Opted Out"
                    a.save()
                except Grub_Student.DoesNotExist:
                    pass
                return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")
            else:
                return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")
        except Grub.DoesNotExist:
            pass
            return HttpResponseRedirect("/ssms/")
    else:
        return HttpResponseRedirect("/ssms/")


@login_required
def user_logout(request):
    logout(request)
    return HttpResponseRedirect('/ssms/')


def import_data(request):
    if request.user.is_superuser:
        done = 0
        if request.method == "POST":
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                def choice_func(row):
                    a = row[0].upper()
                    if a[4] == "H" or a[4] == "P":
                        if (int(a[0:4]) < 2017):
                            b = a[4] + a[0:4] + a[9:12]  # "P" and "0" Not included
                        else:
                            b = a[4] + a[0:4] + a[8:12]  # "P" Not included , 0 included
                    else:
                        if (int(a[0:4]) < 2017):
                            b = 'f' + a[0:4] + a[9:12]  # "P" and "0" Not included
                        else:
                            b = 'f' + a[0:4] + a[8:12]  # "P" Not included , 0 included
                    row.append(b.lower())
                    row[0] = row[0].upper()
                    return row

                request.FILES['file'].save_to_database(
                    model=Student,
                    initializer=choice_func,
                    mapdict=['bits_id', 'name', 'bhawan', 'room_no', 'user_id', ]
                )
                done = 1

        else:
            form = UploadFileForm()
        return render(request, 'ssms/upload_form.html', {'form': form, 'd': done})
    else:
        return HttpResponseRedirect("/ssms/ssms/login/")


def export(request):
    student = Student.objects.all()
    b = []
    for i in student:
        b.append(i.name)
    workbook = xlsxwriter.Workbook('media/' + 'uploaded_student_list.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.set_column('A:A', 20)
    bold = workbook.add_format({'bold': 1})
    worksheet.write('A1', 'User ID', bold)
    row = 1
    col = 0
    for i in b:
        worksheet.write_string(row, col, i)
        row += 1
    workbook.close()
    return (b)
    return HttpResponseRedirect('media/' + 'uploaded_student_list.xlsx')


def allocate(grub, grubstu, no, batchlist):
    lenstu = len(grubstu)
    batchgroup = int(math.ceil(len(grubstu) / no))
    count = 0
    j = 0
    for i in grubstu:
        i.batch = batchlist[j]
        i.save()
        count = count + 1
        if (count > batchgroup):
            j = j + 1
            count = 0
# helper function takes arguments as gmid, request and boolean meal_type. meal_type = true for VegAllocation
# meal_type = false for NonVegAllocation
def batch_allocate(gmid, request, meal_type):
    grub = Grub.objects.get(gm_id=gmid)
    batch_list = [] #generates a batch list to be passed onto allocate(grub, grubstu, no, batchlist)
    name_list = {1: "A", 2: "B", 3: "C", 4: "D"}
    if meal_type:
        #change html page inorder to map VegAllocate1,2,3 in a dict and then access it
        if "VegAllocate1" in request.POST:
            n = 1
        elif "VegAllocate2" in request.POST:
            n = 2
        elif "VegAllocate3" in request.POST:
            n = 3
        else :
            n = 4

    else:
        if "NonVegAllocate1" in request.POST:
            n = 1
        elif "NonVegAllocate2" in request.POST:
            n = 2
        elif "NonVegAllocate3" in request.POST:
            n = 3
        else :
            n = 4

    for i in range(1, n + 1):
        batch = Batch.objects.create(
            gm_id=grub,
            meal="Veg" if meal_type else "Non Veg",
            batch_name=name_list[i],
            color=request.POST["batch{}color".format(i)],
            timing=request.POST["batch{}time".format(i)],
        )
        batch_list.append(batch)

    grubstu = Grub_Student.objects.filter(
        gm_id=grub.gm_id, status="Signed Up").order_by('bhawan')
    allocate(grub, grubstu, n, batch_list) #allocating the batches using the batchlist generated


@user_passes_test(lambda user: user.is_superuser, login_url='ssms/coord/login')
def ssms_grub_batchallocation(request, gmid):

    # redundant veg and non veg references could not be referenced from a dict because of different varaible headers in
    # case of grub.meal = "Both"

    context_dict = {}
    grub = Grub.objects.get(gm_id=gmid)
    e = datechecker(gmid)
    context_dict["e"] = e
    if e == 4 :
        context_dict["error"] = "Grub Over"
        return render(request, 'ssms/batch_allocation.html', context_dict)
    elif e < 3 :
        context_dict["error"] = "Can Allocate Batches only after students cancellation deadline gets over"
        return render(request, 'ssms/batch_allocation.html', context_dict)

    if not request.method == 'POST' :
        context_dict["color"] = ["Pink", "Yellow", "Blue", "Green"]
        context_dict["grub"] = grub
        if grub.meal == "Both":
            both = Both.objects.get(gm_id=grub)
            context_dict["veg"] = both
            context_dict["nonveg"] = both
        context_dict["time"] = ("8:00", "8:15", "8:30", "8:45", "9:00", "9:15",
                                "9:30", "9:45", "10:00", "10:15", "10:30", "10:45", "11:00")
        return render(request, 'ssms/batch_allocation.html', context_dict)


    if grub.meal == "Veg" :
        veg = Veg.objects.get(gm_id=grub)
        batch_allocate(gmid, request, True)# calling batch_allocate with VegAllocate
        veg.batch_allocated = "Yes"
        veg.save()
        grub.batch_allocated = "Yes"
    elif grub.meal == "Non Veg" :
        nonveg = NonVeg.objects.get(gm_id=grub)
        batch_allocate(gmid, request, False)# calling batch_allocate with NonVegAllocate
        nonveg.batch_allocated = "Yes"
        nonveg.save()
        grub.batch_allocated = "Yes"

    else :
        both_meals = Both.objects.get(gm_id=grub)
        if "VegAllocate1" in request.POST or "VegAllocate2" in request.POST or "VegAllocate3" in request.POST or "VegAllocate4" in request.POST:
            batch_allocate(gmid, request, True) # calling batch_allocate once with VegAllocate
            both_meals.veg_batch_allocated = "Yes"
        elif "NonVegAllocate1" in request.POST or "NonVegAllocate2" in request.POST or "NonVegAllocate3" in request.POST or "NonVegAllocate4" in request.POST:
            batch_allocate(gmid, request, False) # calling batch_allocate once with NonVegAllocate
            both_meals.nonveg_batch_allocated = "Yes"
        both_meals.save()
        context_dict["veg"] = both_meals
        context_dict["nonveg"] = both_meals
        if both_meals.veg_batch_allocated == "Yes" and both_meals.nonveg_batch_allocated == "Yes":
            grub.batch_allocated = "Yes"

    grub.save()
    if grub.batch_allocated == "Yes":
        return HttpResponseRedirect("/ssms/stats/" + gmid)
    context_dict["color"] = ["Pink", "Yellow", "Blue", "Green"]
    context_dict["grub"] = grub
    context_dict["time"] = ("8:00", "8:15", "8:30", "8:45", "9:00", "9:15",
                            "9:30", "9:45", "10:00", "10:15", "10:30", "10:45", "11:00")
    return render(request, 'ssms/batch_allocation.html', context_dict)




def student_grub_feedback(request, gmid):
    if request.user.is_authenticated():  # and not request.user.is_staff:
        try:
            grub = Grub.objects.get(gm_id=gmid, status="Active")
            d = datechecker(gmid)
            print("here")
            if d == 4:
                print("here")
                try:
                    a = Grub_Student.objects.get(
                        gm_id=gmid, user_id=str(request.user), status="Signed Up")
                    if request.method == 'POST':
                        fb_form = FeedbackForm(request.POST)
                        if fb_form.is_valid():
                            feedb_form = fb_form.save(commit=False)
                            feedb_form.user = request.user
                            feedb_form.gm_id = grub
                            feedb_form.stugm_id = a.student_id
                            feedb_form.meal_type = a.meal
                            feedb_form.save()
                            a.feedback_given = 1
                            a.save()
                            return HttpResponse("here")
                        else:
                            return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")
                    else:
                        fb_form = FeedbackForm()
                        return render(request, 'ssms/createfb.html', {'grub': grub, 'fb_form': fb_form, "grubstu": a})
                except Grub_Student.DoesNotExist:
                    return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")
            else:
                return HttpResponseRedirect("/ssms/student/grub/" + gmid + "/")
        except Grub.DoesNotExist:
            pass
            return HttpResponseRedirect("/ssms/")
    else:
        return HttpResponseRedirect("/ssms/")


def menu_upload(request):
    # check if request is post and handle appropriately by storing file in media
    if request.user.is_superuser:
        context_dict = {}
        if request.method == 'POST' and request.FILES['myfile']:
            # file path can be filled manually or using file url generated by file upload.
            try:
                name = request.FILES['myfile'].name
                wb = openpyxl.load_workbook(request.FILES['myfile'])
                sheet = wb.active
                max_val = 32
                lunchat = 0  # position of lunch field in excel
                dinnerat = 0  # position of dinner field in excel
                countcell = 0  # sheet.max_row
                for col in sheet.iter_cols():
                    countcell = 0
                    for i in col:
                        countcell += 1
                        if (str(i.value).lower().startswith("lunch")):
                            lunchat = i.row - 1  # 0-indexing
                        if (str(i.value).lower().startswith("dinner")):
                            dinnerat = i.row - 1  # 0-indexing
                    if (lunchat != 0 and dinnerat != 0):
                        break
                breakfast = lunch = dinner = ""  # fields of menu model
                lunchgrub = dinnergrub = "0"  # fields of menu model
                for col in sheet.iter_cols():
                    breakfast = lunch = dinner = ""  # fields of menu model
                    lunchgrub = dinnergrub = "0"  # fields of menu model
                    date, day = col[0].value, col[1].value
                    for j in range(2, lunchat):
                        if col[j].value:
                            breakfast += str(col[j].value) + "###"
                    for j in range(lunchat, dinnerat):
                        if (str(col[j].value).lower().startswith("grub")):
                            lunchgrub = "1"
                        elif (str(col[j].value).lower().startswith("lunch")):
                            continue
                        elif col[j].value:
                            lunch += str(col[j].value) + "###"
                    for j in range(dinnerat, countcell):
                        if (str(col[j].value).lower().startswith("grub")):
                            dinnergrub = "1"
                        elif (str(col[j].value).lower().startswith("dinner")):
                            continue
                        elif col[j].value:
                            dinner += str(col[j].value) + "###"
                    try:
                        meal = Meal.objects.get(date=date)
                        meal.day = day
                        meal.lunch = lunch
                        meal.dinner = dinner
                        meal.breakfast = breakfast
                        meal.lunchgrub = lunchgrub
                        meal.dinnergrub = dinnergrub
                        meal.save()
                    except:
                        Meal.objects.create(date=date, day=day, breakfast=breakfast,
                                            dinner=dinner, lunch=lunch, lunchgrub=lunchgrub, dinnergrub=dinnergrub)

                context_dict["error"] = "Succesfully Uploaded!!"
                return render(request, 'ssms/menu_upload.html', context_dict)
            except Exception as e:
                print(e)
                context_dict["error"] = e
                return render(request, 'ssms/menu_upload.html', context_dict)
        else:
            context_dict["error"] = "No file uploaded"
            return render(request, 'ssms/menu_upload.html', context_dict)
    else:
        return HttpResponseRedirect("/ssms/ssms/login/")


def menu_display(request):
    # if database has no data then catch error and display menu.html page
    current_date = date.today().isoformat()
    display_date = date.strftime(date.today(), "%a-%d-%b")
    try:
        meal = Meal.objects.get(date=current_date)
        day = meal.day
        breakfast = str(meal.breakfast).split("###")[:-1]  # to cut the last "###"
        lunch = str(meal.lunch).split("###")[:-1]  # to cut the last "###"
        dinner = str(meal.dinner).split("###")[:-1]  # to cut the last "###"
        lunchgrub = meal.lunchgrub
        dinnergrub = meal.dinnergrub
        return render(request, 'ssms/menu.html', {'day': display_date, 'breakfast': breakfast, 'lunch': lunch,
                                                  'dinner': dinner, 'lunchgrub': lunchgrub, 'dinnergrub': dinnergrub})
    except:
        error = "Not Updated :("
        return render(request, 'ssms/menu.html', {'day': display_date, 'error': error})
